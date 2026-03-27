"""
Dashboard — Organisations durabilité & ODD
IQ · FSTQ · Desjardins Capital · Développement Économique Canada
"""
import io
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Organisations — Durabilité & ODD",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.html("""
<style>
  [data-testid="stAppViewContainer"] { background: #0B1120; }
  [data-testid="stHeader"]           { background: transparent; }
  section[data-testid="stSidebar"]   { background: #111827; }
  .block-container { padding-top:1.2rem; padding-bottom:2rem; max-width:1500px; }

  .page-title   { font-size:1.6rem; font-weight:900; color:#F9FAFB; letter-spacing:.04em; }
  .page-sub     { font-size:.75rem; color:#6B7280; margin-top:3px; }
  .section-hdr  {
    font-size:.6rem; letter-spacing:.2em; text-transform:uppercase;
    color:#9CA3AF; background:#1F2937; border-radius:4px;
    padding:3px 10px; display:inline-block; margin-bottom:8px;
  }
  .org-card {
    background:#111827; border:1px solid #1F2937; border-radius:8px;
    padding:14px 16px; margin-bottom:10px;
  }
  .org-name  { font-size:1rem; font-weight:800; color:#34D399; }
  .org-type  { font-size:.7rem; color:#9CA3AF; margin-top:2px; }
  .org-meta  { font-size:.75rem; color:#D1D5DB; margin-top:6px; }
  .org-desc  { font-size:.72rem; color:#9CA3AF; margin-top:6px; line-height:1.5; }

  .badge-e   { background:rgba(16,185,129,.15); color:#34D399;
               border:1px solid rgba(16,185,129,.3); border-radius:4px;
               padding:2px 8px; font-size:.62rem; font-weight:700;
               display:inline-block; margin:2px; }
  .badge-s   { background:rgba(59,130,246,.15); color:#93C5FD;
               border:1px solid rgba(59,130,246,.3); border-radius:4px;
               padding:2px 8px; font-size:.62rem; font-weight:700;
               display:inline-block; margin:2px; }
  .badge-g   { background:rgba(245,158,11,.15); color:#FCD34D;
               border:1px solid rgba(245,158,11,.3); border-radius:4px;
               padding:2px 8px; font-size:.62rem; font-weight:700;
               display:inline-block; margin:2px; }
  .badge-odd { background:#1E3A5F; color:#93C5FD;
               border:1px solid #2D4A7A; border-radius:12px;
               padding:2px 10px; font-size:.62rem; font-weight:600;
               display:inline-block; margin:2px; }
  .badge-odd-direct { background:rgba(16,185,129,.2); color:#34D399;
               border:1px solid rgba(16,185,129,.4); border-radius:12px;
               padding:2px 10px; font-size:.62rem; font-weight:700;
               display:inline-block; margin:2px; }
  .neq-box {
    background:#0F172A; border:1px solid #1E3A5F; border-radius:6px;
    padding:6px 12px; font-size:.7rem; color:#94A3B8; margin-top:6px;
  }
  .divider { border-top:1px solid #1F2937; margin:12px 0; }
  .note-box {
    background:#1C1917; border:1px solid #78350F; border-radius:6px;
    padding:8px 14px; font-size:.7rem; color:#D97706; margin-top:8px;
  }
  .kpi-val  { font-size:1.4rem; font-weight:900; color:#34D399; }
  .kpi-lbl  { font-size:.68rem; color:#6B7280; margin-top:2px; }
</style>
""")

def _hex_rgba(hex_color: str, alpha: float) -> str:
    """Convertit '#RRGGBB' en 'rgba(R,G,B,alpha)' pour plotly."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"


@st.cache_data
def _build_excel() -> bytes:
    """Génère le fichier Excel en mémoire (compatible Streamlit Cloud)."""
    C_HEADER="1E3A5F"; C_ACCENT="10B981"; C_WH="FFFFFF"; C_DK="1F2937"
    C_GRAY="6B7280"; C_R1="EFF6FF"; C_R2="F0FDF9"; C_R3="FFF7ED"; C_R4="F5F3FF"
    C_YES="D1FAE5"; C_IND="FEF3C7"; C_NO="F3F4F6"
    def fl(h): return PatternFill("solid", fgColor=h)
    def bd():
        s=Side(style="thin",color="D1D5DB")
        return Border(left=s,right=s,top=s,bottom=s)
    def ct(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def lf(): return Alignment(horizontal="left",vertical="center",wrap_text=True)
    def hf(sz=11,bold=True,color=None): return Font(name="Arial",bold=bold,size=sz,color=color or C_WH)
    def bf(sz=10,color=None): return Font(name="Arial",size=sz,color=color or C_DK)

    wb = Workbook()
    ws1 = wb.active; ws1.title="Profils organisations"; ws1.sheet_view.showGridLines=False
    ws1.merge_cells("A1:J1"); ws1["A1"]="ORGANISATIONS - DURABILITE & ODD | IQ - FSTQ - Desjardins Capital - DEC"
    ws1["A1"].font=hf(14); ws1["A1"].fill=fl(C_HEADER); ws1["A1"].alignment=ct(); ws1.row_dimensions[1].height=32
    ws1.merge_cells("A2:J2"); ws1["A2"]="Sources : REQ - Sites officiels - Rapports ESG 2024-2025 | 2026-03-27"
    ws1["A2"].font=Font(name="Arial",italic=True,size=8,color=C_GRAY); ws1["A2"].fill=fl("F9FAFB"); ws1["A2"].alignment=ct(); ws1.row_dimensions[2].height=16
    hdrs=["Organisation","Nom officiel","NEQ","Adresse","Ville/CP","Type","Fond.","Actif gere","Site web","Mission"]
    wdts=[22,38,15,42,20,30,8,24,28,70]
    for i,(h,w) in enumerate(zip(hdrs,wdts),1):
        c=ws1.cell(3,i,h); c.font=hf(10); c.fill=fl(C_ACCENT); c.alignment=ct(); c.border=bd()
        ws1.column_dimensions[get_column_letter(i)].width=w
    ws1.row_dimensions[3].height=28
    rows_data=[
        ("Investissement Quebec (IQ)","Investissement Quebec","A verifier - REQ","1195, av. Lavigerie, bur. 060","Quebec G1V 4N3","Societe d'Etat (gouv. QC)",2011,"7,5 G$ (2024)","investquebec.com","Contribuer au developpement economique du Quebec via prets, capital-actions, garanties et credits d'impot. Plan DD 2023-2028 aligne Agenda 2030. Questionnaire ESG obligatoire. Signataire Finance Montreal. Integration TCFD/GIFCC."),
        ("Fonds de solidarite FTQ (FSTQ)","Fonds de solidarite des travailleurs du Quebec (F.T.Q.)","A verifier - REQ","545, boul. Cremazie Est, bur. 200","Montreal H2M 2W4","Fonds capital developpement (loi speciale QC)",1983,"21,9 G$ (mai 2025)","fondsftq.com","Fonds de capital de developpement pour creer et proteger des emplois. Vision 2022-2027 : prosperite durable et inclusive. Cible 12 G$ en actifs durables (9+ G$ atteints). Siege LEED v5 Platine 2025. Cadre 6 rendements societaux."),
        ("Desjardins Capital (CRCD)","Capital regional et cooperatif Desjardins (CRCD)","A verifier - REQ","2 Complexe Desjardins, bur. 1717","Montreal H5B 1B2","Fonds investissement capital developpement",2001,"2,7 G$ CRCD / ~4,9 G$ total","capitalregional.com","75 % des PME hors grands centres. Accompagnement ESG, releve et numerique. Desjardins : zero emission nette 2040, 6 G$+ transition energetique, SBTi valide, rapports GRI/SASB/PRB."),
        ("Dev. Economique Canada (DEC)","Developpement economique Canada pour les regions du Quebec","S.O. - Agence federale","800, boul. Rene-Levesque O., bur. 500","Montreal H3B 1X9","Agence federale (gouv. Canada)",1991,"316,2 M$ (2024-2025)","dec.canada.ca","Promouvoir le developpement economique des regions du Quebec. 65,4 M$ en projets verts 2024-2025. Inclusion autochtone : 19,47 % valeur contractuelle. Strategie DD alignee ODD 8/9/10/11/12/13."),
    ]
    rcs=[C_R1,C_R2,C_R3,C_R4]
    for ri,(row,rc) in enumerate(zip(rows_data,rcs),4):
        ws1.row_dimensions[ri].height=90
        for ci,val in enumerate(row,1):
            c=ws1.cell(ri,ci,val); c.fill=fl(rc); c.border=bd()
            if ci==1: c.font=Font(name="Arial",bold=True,size=10,color="1E3A5F"); c.alignment=lf()
            elif ci==10: c.font=bf(9); c.alignment=lf()
            elif ci==7: c.font=bf(); c.alignment=ct()
            else: c.font=bf(); c.alignment=lf()
    ws1.merge_cells("A8:J8"); ws1["A8"]="* NEQ : registreentreprises.gouv.qc.ca. DEC (federale) non assujettie au REQ provincial."
    ws1["A8"].font=Font(name="Arial",italic=True,size=8,color=C_GRAY); ws1["A8"].fill=fl(C_R3); ws1["A8"].alignment=lf(); ws1.row_dimensions[8].height=20
    ws1.freeze_panes="A4"

    ws2=wb.create_sheet("Matrice ODD"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:F1"); ws2["A1"]="MATRICE D'ALIGNEMENT ODD / SDG"
    ws2["A1"].font=hf(14); ws2["A1"].fill=fl(C_HEADER); ws2["A1"].alignment=ct(); ws2.row_dimensions[1].height=30
    ws2.merge_cells("A2:F2"); ws2["A2"]="OK = Confirme | (indirect) = Indirect/sectoriel | --- = Non documente"
    ws2["A2"].font=Font(name="Arial",italic=True,size=9,color=C_GRAY); ws2["A2"].fill=fl("F9FAFB"); ws2["A2"].alignment=ct(); ws2.row_dimensions[2].height=16
    oh2=["ODD","Theme","IQ","FSTQ","Desjardins Capital","DEC"]
    cw2=[9,32,22,22,22,24]
    for i,(h,w) in enumerate(zip(oh2,cw2),1):
        c=ws2.cell(3,i,h); c.fill=fl(C_ACCENT if i>2 else C_HEADER); c.font=hf(10); c.alignment=ct(); c.border=bd()
        ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.row_dimensions[3].height=36
    odd_rows=[
        ("ODD 1","Pas de pauvrete","E53E3E",["(indirect)","OK","(indirect)","---"]),
        ("ODD 3","Bonne sante","38A169",["---","OK","---","(indirect)"]),
        ("ODD 4","Education","D69E2E",["---","OK","---","---"]),
        ("ODD 5","Egalite sexes","D53F8C",["(indirect)","OK","(indirect)","OK"]),
        ("ODD 6","Eau propre","3182CE",["---","OK","---","---"]),
        ("ODD 7","Energie propre","F6AD55",["OK","OK","OK","OK"]),
        ("ODD 8","Travail decent","744210",["OK","OK","OK","OK"]),
        ("ODD 9","Industrie & innovation","C05621",["OK","OK","OK","OK"]),
        ("ODD 10","Inegalites reduites","DD6B20",["(indirect)","(indirect)","(indirect)","OK"]),
        ("ODD 11","Villes durables","7B341E",["(indirect)","OK","OK","OK"]),
        ("ODD 12","Conso. responsable","276749",["---","---","(indirect)","OK"]),
        ("ODD 13","Action climatique","22543D",["OK","OK","OK","OK"]),
        ("ODD 15","Vie terrestre","1A4731",["(indirect)","OK","(indirect)","---"]),
        ("ODD 17","Partenariats","1A365D",["OK","---","OK","(indirect)"]),
    ]
    for ri,(code,label,color,vals) in enumerate(odd_rows,4):
        ws2.row_dimensions[ri].height=24
        c1=ws2.cell(ri,1,code); c1.font=Font(name="Arial",bold=True,size=10,color=C_WH); c1.fill=fl(color); c1.alignment=ct(); c1.border=bd()
        c2=ws2.cell(ri,2,label); c2.font=bf(10); c2.fill=fl("F9FAFB"); c2.alignment=lf(); c2.border=bd()
        for ci,v in enumerate(vals,3):
            cell=ws2.cell(ri,ci,v)
            if v=="OK": cell.fill=fl(C_YES); cell.font=Font(name="Arial",bold=True,size=10,color="065F46")
            elif "(indirect)" in v: cell.fill=fl(C_IND); cell.font=Font(name="Arial",bold=True,size=10,color="92400E")
            else: cell.fill=fl(C_NO); cell.font=Font(name="Arial",size=10,color="9CA3AF")
            cell.alignment=ct(); cell.border=bd()
    tr=len(odd_rows)+4; ws2.row_dimensions[tr].height=26
    for c in [1,2]:
        cell=ws2.cell(tr,c,"TOTAL ODD confirmes" if c==2 else "TOTAL"); cell.font=hf(10); cell.fill=fl(C_HEADER); cell.alignment=ct(); cell.border=bd()
    for i,t in enumerate([7,11,7,6],3):
        cell=ws2.cell(tr,i,t); cell.font=Font(name="Arial",bold=True,size=13,color=C_WH); cell.fill=fl(C_ACCENT); cell.alignment=ct(); cell.border=bd()
    ws2.freeze_panes="C4"

    ws3=wb.create_sheet("Engagements ESG"); ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:G1"); ws3["A1"]="ENGAGEMENTS ESG & DURABILITE"
    ws3["A1"].font=hf(14); ws3["A1"].fill=fl(C_HEADER); ws3["A1"].alignment=ct(); ws3.row_dimensions[1].height=30
    eh=["Organisation","Categorie","Engagement","Cible","Statut 2024-2025","Source","URL"]
    ew=[24,16,55,32,22,28,30]
    for i,(h,w) in enumerate(zip(eh,ew),1):
        c=ws3.cell(2,i,h); c.font=hf(10); c.fill=fl(C_ACCENT); c.alignment=ct(); c.border=bd()
        ws3.column_dimensions[get_column_letter(i)].width=w
    ws3.row_dimensions[2].height=26
    cf={"Environnement":"D1FAE5","Social":"DBEAFE","Gouvernance":"FEF3C7"}
    om={"Investissement Quebec":C_R1,"FSTQ":C_R2,"Desjardins Capital":C_R3,"DEC":C_R4}
    esg=[
        ["Investissement Quebec","Gouvernance","Questionnaire ESG obligatoire","100 % dossiers","En vigueur","Rapport ESG IQ 2024","investquebec.com"],
        ["Investissement Quebec","Gouvernance","Signataire Finance Montreal","---","Confirme 2024","Finance Montreal","finance-montreal.com"],
        ["Investissement Quebec","Environnement","Integration TCFD/GIFCC","---","En cours","PADD 2023-2028","investquebec.com"],
        ["Investissement Quebec","Environnement","Reduction intensite carbone","Cible 2028","En cours","PADD 2023-2028","investquebec.com"],
        ["Investissement Quebec","Social","Plan DD 2023-2028 (Agenda 2030)","ODD Quebec","Actif","Gouv. QC / IQ","investquebec.com"],
        ["FSTQ","Environnement","Actifs de developpement durable","12 G$ d'ici 2027","9+ G$ atteints","Rapport DD FTQ 2024","fondsftq.com"],
        ["FSTQ","Social","Actions a impact employes","100 000 d'ici 2027","69 000+ realisees","Rapport DD FTQ 2024","fondsftq.com"],
        ["FSTQ","Social","Actionnaires sans regime retraite","100 000 d'ici 2027","En cours","Rapport DD FTQ 2024","fondsftq.com"],
        ["FSTQ","Gouvernance","Rapport lutte travail force (Loi S-211)","Annuel","Publie 2024","Gouv. Canada","fondsftq.com"],
        ["FSTQ","Environnement","Siege LEED v5 O+M Platine / BOMA BEST","---","Certifie 2025","LEED / BOMA","fondsftq.com"],
        ["FSTQ","Gouvernance","Cadre 6 rendements societaux","---","Actif","Cadre FTQ","fondsftq.com"],
        ["Desjardins Capital","Environnement","Zero emission nette Desjardins","2040","En cours","Desj. ESG 2024","desjardins.com"],
        ["Desjardins Capital","Environnement","Energies renouvelables","2 G$ engages","Atteint avant 2025","Desj. ESG 2024","desjardins.com"],
        ["Desjardins Capital","Environnement","Transition energetique","6 G$+ depuis 2020","En cours","Desj. ESG 2024","desjardins.com"],
        ["Desjardins Capital","Gouvernance","Ambition SBTi (1,5 C)","Validee","Confirmee","SBTi/Desjardins","desjardins.com"],
        ["Desjardins Capital","Gouvernance","Rapports GRI / SASB / PRB / PSI","Annuel","Publie 2024","Desj. ESG","desjardins.com"],
        ["Desjardins Capital","Social","Integration ESG + releve + numerique PME","100 % portefeuille","En cours","Desj. Capital","capitalregional.com"],
        ["DEC","Environnement","Investissements technologies propres","25 M$/an","65,4 M$ en 2024-25","Strat. DD DEC","dec.canada.ca"],
        ["DEC","Environnement","Flotte ZEV ou hybride","35 % mars 2027","En cours","Strat. DD DEC","dec.canada.ca"],
        ["DEC","Environnement","Reduction GES flotte","-15 % 2026-27","En cours","Strat. DD DEC","dec.canada.ca"],
        ["DEC","Social","Inclusion autochtone","19,47 % valeur contractuelle","Atteint 2024-25","Rapport DEC","dec.canada.ca"],
        ["DEC","Gouvernance","Strategie ministerielle DD 2024-2025","ODD 8/9/10/11/12/13","Publiee 2024","Gouv. Canada","dec.canada.ca"],
    ]
    for ri,row in enumerate(esg,3):
        ws3.row_dimensions[ri].height=42; rc=om.get(row[0],"FFFFFF")
        for ci,val in enumerate(row,1):
            cell=ws3.cell(ri,ci,val); cell.border=bd()
            if ci==1: cell.font=Font(name="Arial",bold=True,size=9,color="1E3A5F"); cell.fill=fl(rc); cell.alignment=lf()
            elif ci==2: cell.font=bf(9); cell.fill=fl(cf.get(val,"F9FAFB")); cell.alignment=ct()
            elif ci in [4,5]: cell.font=bf(9); cell.fill=fl(rc); cell.alignment=ct()
            else: cell.font=bf(9); cell.fill=fl(rc); cell.alignment=lf()
    ws3.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ── Données ───────────────────────────────────────────────────────────────────
ORGS = [
    {
        "id": "IQ",
        "nom": "Investissement Québec",
        "court": "IQ",
        "type": "Société d'État · Gouvernement du Québec",
        "fondation": 2011,
        "ville": "Québec (QC)",
        "adresse": "1195, av. Lavigerie, bur. 060, Québec G1V 4N3",
        "aum": "7,5 G$",
        "neq": "À vérifier — registreentreprises.gouv.qc.ca",
        "web": "investquebec.com",
        "couleur": "#3B82F6",
        "mission": (
            "Contribuer au développement économique du Québec en stimulant l'innovation, "
            "l'entrepreneuriat et la croissance de l'investissement et des exportations. "
            "Soutien via prêts, capital-actions, garanties et crédits d'impôt dans toutes les régions."
        ),
        "esg": {
            "E": ["Intégration TCFD/GIFCC", "Réduction intensité carbone portefeuille"],
            "S": ["Plan DD 2023-2028 (Agenda 2030)", "Présence dans toutes les régions QC"],
            "G": ["Questionnaire ESG obligatoire 100 % des dossiers", "Signataire Finance Montréal"],
        },
        "odds_direct": ["ODD 7", "ODD 8", "ODD 9", "ODD 13", "ODD 17"],
        "odds_indirect": ["ODD 1", "ODD 5", "ODD 10", "ODD 11", "ODD 15"],
    },
    {
        "id": "FSTQ",
        "nom": "Fonds de solidarité FTQ",
        "court": "FSTQ",
        "type": "Fonds de capital de développement · Loi spéciale, Assemblée nationale QC",
        "fondation": 1983,
        "ville": "Montréal (QC)",
        "adresse": "545, boul. Crémazie Est, bur. 200, Montréal H2M 2W4",
        "aum": "21,9 G$",
        "neq": "À vérifier — registreentreprises.gouv.qc.ca",
        "web": "fondsftq.com",
        "couleur": "#10B981",
        "mission": (
            "Fonds de capital de développement (loi spéciale) pour créer, maintenir et protéger "
            "des emplois via des investissements dans des PME de toutes les régions. "
            "Vision 2022-2027 : prospérité durable, responsable et inclusive."
        ),
        "esg": {
            "E": ["Cible 12 G$ actifs durables d'ici 2027 (9+ G$ atteints)", "Siège LEED v5 O+M Platine 2025"],
            "S": ["100 000 actions à impact employés (69 000+ réalisées)", "100 000 nouveaux actionnaires sans régime retraite"],
            "G": ["Rapport annuel lutte travail forcé (Loi S-211)", "Cadre des 6 rendements sociétaux"],
        },
        "odds_direct": ["ODD 1", "ODD 3", "ODD 4", "ODD 5", "ODD 6", "ODD 7", "ODD 8", "ODD 9", "ODD 11", "ODD 13", "ODD 15"],
        "odds_indirect": ["ODD 10"],
    },
    {
        "id": "CRCD",
        "nom": "Desjardins Capital",
        "court": "Desjardins Capital (CRCD)",
        "type": "Fonds d'investissement en capital de développement · Société publique à actionnaires",
        "fondation": 2001,
        "ville": "Montréal (QC)",
        "adresse": "2 Complexe Desjardins, bur. 1717, Montréal",
        "aum": "2,7 G$ CRCD / ~4,9 G$ total",
        "neq": "À vérifier — registreentreprises.gouv.qc.ca",
        "web": "capitalregional.com",
        "couleur": "#F59E0B",
        "mission": (
            "Injecter des capitaux dans des coopératives et des PME de toutes les régions "
            "du Québec (75 % hors Montréal/Québec). Accompagnement relève entrepreneuriale, "
            "numérique et ESG. Écosystème Desjardins vise zéro émission nette 2040."
        ),
        "esg": {
            "E": ["Zéro émission nette Desjardins 2040", "6 G$+ transition énergétique depuis 2020", "2 G$ énergies renouvelables (atteint)"],
            "S": ["Intégration ESG + accompagnement relève et numérique PME", "75 % PME hors grands centres"],
            "G": ["Ambition SBTi validée (1,5°C)", "Rapports GRI / SASB / PRB / PSI"],
        },
        "odds_direct": ["ODD 7", "ODD 8", "ODD 9", "ODD 11", "ODD 13", "ODD 17"],
        "odds_indirect": ["ODD 1", "ODD 5", "ODD 10", "ODD 12", "ODD 15"],
    },
    {
        "id": "DEC",
        "nom": "Développement économique Canada",
        "court": "DEC",
        "type": "Agence fédérale · Gouvernement du Canada · 11 bureaux régionaux au Québec",
        "fondation": 1991,
        "ville": "Montréal (QC)",
        "adresse": "800, boul. René-Lévesque O., bur. 500, Montréal H3B 1X9",
        "aum": "316,2 M$ (budget 2024-2025)",
        "neq": "S.O. — Agence fédérale (non assujettie au REQ provincial)",
        "web": "dec.canada.ca",
        "couleur": "#8B5CF6",
        "mission": (
            "Promouvoir le développement économique à long terme des régions du Québec, "
            "notamment là où la croissance est lente. Finance PME, OBNL et 67 SADC/CAE. "
            "65,4 M$ investis en projets verts en 2024-2025."
        ),
        "esg": {
            "E": ["65,4 M$ en projets verts 2024-2025", "Flotte 35 % zéro émission d'ici 2027", "Réduction GES flotte −15 % d'ici 2026-2027"],
            "S": ["Inclusion autochtone : 19,47 % valeur contractuelle 2024-2025", "Entrepreneuriat inclusif : jeunes, femmes, autochtones"],
            "G": ["Stratégie ministerielle DD 2024-2025", "Alignement ODD 8/9/10/11/12/13"],
        },
        "odds_direct": ["ODD 5", "ODD 8", "ODD 9", "ODD 10", "ODD 11", "ODD 12", "ODD 13"],
        "odds_indirect": ["ODD 3", "ODD 7", "ODD 17"],
    },
]

ODD_LABELS = {
    "ODD 1": "Pas de pauvreté", "ODD 2": "Faim zéro", "ODD 3": "Bonne santé",
    "ODD 4": "Éducation de qualité", "ODD 5": "Égalité des sexes",
    "ODD 6": "Eau propre", "ODD 7": "Énergie propre",
    "ODD 8": "Travail décent", "ODD 9": "Industrie & innovation",
    "ODD 10": "Inégalités réduites", "ODD 11": "Villes durables",
    "ODD 12": "Conso. responsable", "ODD 13": "Action climatique",
    "ODD 14": "Vie aquatique", "ODD 15": "Vie terrestre",
    "ODD 16": "Paix & justice", "ODD 17": "Partenariats",
}

ODD_COLORS = {
    "ODD 1": "#E53E3E", "ODD 3": "#38A169", "ODD 4": "#D69E2E",
    "ODD 5": "#D53F8C", "ODD 6": "#3182CE", "ODD 7": "#F6AD55",
    "ODD 8": "#744210", "ODD 9": "#C05621", "ODD 10": "#DD6B20",
    "ODD 11": "#7B341E", "ODD 12": "#276749", "ODD 13": "#22543D",
    "ODD 15": "#1A4731", "ODD 17": "#1A365D",
}

# ── En-tête ───────────────────────────────────────────────────────────────────
c1, c2 = st.columns([3, 1])
with c1:
    st.html('<div class="page-title">🌿 Organisations — Durabilité & ODD</div>')
    st.html('<div class="page-sub">IQ &nbsp;·&nbsp; FSTQ &nbsp;·&nbsp; Desjardins Capital &nbsp;·&nbsp; Développement Économique Canada &nbsp;|&nbsp; Mis à jour : 2026-03-27</div>')
with c2:
    st.download_button(
        label="⬇ Télécharger Excel",
        data=_build_excel(),
        file_name="Organisations_Durabilite_ODD.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.html('<div class="divider"></div>')

# ── KPIs globaux ──────────────────────────────────────────────────────────────
st.html('<span class="section-hdr">VUE D\'ENSEMBLE</span>')
k1, k2, k3, k4, k5 = st.columns(5)

all_odds_direct = set()
for o in ORGS:
    all_odds_direct.update(o["odds_direct"])

with k1:
    st.html('<div class="org-card"><div class="kpi-val">4</div><div class="kpi-lbl">Organisations</div></div>')
with k2:
    st.html('<div class="org-card"><div class="kpi-val">~33 G$</div><div class="kpi-lbl">Actifs gérés totaux</div></div>')
with k3:
    st.html(f'<div class="org-card"><div class="kpi-val">{len(all_odds_direct)}</div><div class="kpi-lbl">ODD couverts (alignement direct)</div></div>')
with k4:
    st.html('<div class="org-card"><div class="kpi-val">3</div><div class="kpi-lbl">Niveaux de gouvernance (munic. / prov. / féd.)</div></div>')
with k5:
    st.html('<div class="org-card"><div class="kpi-val">4</div><div class="kpi-lbl">Rapports ESG/DD publiés (2024)</div></div>')

st.html('<div class="divider"></div>')

# ── Onglets ───────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(["🏢 Profils", "🎯 Matrice ODD", "📊 Comparatif ESG", "📋 Données brutes"])

# ══ TAB 1 — Profils ══════════════════════════════════════════════════════════
with tab1:
    for org in ORGS:
        with st.expander(f"**{org['nom']}** ({org['court']}) — {org['ville']}", expanded=True):
            left, right = st.columns([3, 2])

            with left:
                st.html(f'<div class="org-name" style="color:{org["couleur"]}">{org["nom"]}</div>')
                st.html(f'<div class="org-type">{org["type"]}</div>')
                st.html(f'<div class="org-meta">📍 {org["adresse"]}</div>')
                st.html(f'<div class="org-meta">🌐 {org["web"]} &nbsp;·&nbsp; 📅 Fondé en {org["fondation"]} &nbsp;·&nbsp; 💰 {org["aum"]}</div>')
                st.html(f'<div class="neq-box">🔢 <b>NEQ :</b> {org["neq"]}</div>')
                st.html(f'<div class="org-desc">{org["mission"]}</div>')

                st.html('<div style="margin-top:10px"><b style="color:#9CA3AF;font-size:.7rem">ENGAGEMENTS ESG</b></div>')
                badges_e = "".join(f'<span class="badge-e">🌱 {x}</span>' for x in org["esg"]["E"])
                badges_s = "".join(f'<span class="badge-s">👥 {x}</span>' for x in org["esg"]["S"])
                badges_g = "".join(f'<span class="badge-g">⚖️ {x}</span>' for x in org["esg"]["G"])
                st.html(f'<div style="margin-top:4px">{badges_e}{badges_s}{badges_g}</div>')

            with right:
                st.html('<div style="margin-bottom:6px"><b style="color:#9CA3AF;font-size:.7rem">ODD ALIGNÉS</b></div>')
                direct_html = "".join(
                    f'<span class="badge-odd-direct" title="{ODD_LABELS.get(o,"")}">✓ {o}</span>'
                    for o in org["odds_direct"]
                )
                indirect_html = "".join(
                    f'<span class="badge-odd" title="{ODD_LABELS.get(o,"")}">◐ {o}</span>'
                    for o in org.get("odds_indirect", [])
                )
                st.html(f'<div>{direct_html}</div><div style="margin-top:4px">{indirect_html}</div>')
                st.html('<div style="margin-top:6px;font-size:.62rem;color:#6B7280">✓ Confirmé &nbsp;·&nbsp; ◐ Indirect</div>')

                # Mini radar ESG
                categories = ["Environnement", "Social", "Gouvernance"]
                scores = [len(org["esg"]["E"]), len(org["esg"]["S"]), len(org["esg"]["G"])]
                fig_r = go.Figure(go.Scatterpolar(
                    r=scores + [scores[0]],
                    theta=categories + [categories[0]],
                    fill="toself",
                    line_color=org["couleur"],
                    fillcolor=_hex_rgba(org["couleur"], 0.2),
                ))
                fig_r.update_layout(
                    polar=dict(
                        bgcolor="#0B1120",
                        radialaxis=dict(visible=True, range=[0, 5], color="#6B7280", gridcolor="#1F2937"),
                        angularaxis=dict(color="#9CA3AF", gridcolor="#1F2937"),
                    ),
                    paper_bgcolor="#111827",
                    plot_bgcolor="#111827",
                    margin=dict(l=20, r=20, t=20, b=20),
                    height=200,
                    showlegend=False,
                )
                st.plotly_chart(fig_r, use_container_width=True, key=f"radar_{org['id']}")

# ══ TAB 2 — Matrice ODD ══════════════════════════════════════════════════════
with tab2:
    st.html('<span class="section-hdr">MATRICE D\'ALIGNEMENT ODD × ORGANISATIONS</span>')

    all_odds = sorted(set(
        o for org in ORGS for o in org["odds_direct"] + org.get("odds_indirect", [])
    ), key=lambda x: int(x.split()[1]))

    rows = []
    for odd in all_odds:
        row = {"ODD": odd, "Thème": ODD_LABELS.get(odd, "")}
        for org in ORGS:
            if odd in org["odds_direct"]:
                row[org["court"]] = 2   # direct
            elif odd in org.get("odds_indirect", []):
                row[org["court"]] = 1   # indirect
            else:
                row[org["court"]] = 0
        rows.append(row)

    df_matrix = pd.DataFrame(rows)
    org_cols = [o["court"] for o in ORGS]

    # Heatmap
    z_vals = df_matrix[org_cols].values
    y_labels = [f"{r['ODD']} — {r['Thème']}" for _, r in df_matrix.iterrows()]
    x_labels = [o["nom"] for o in ORGS]

    colorscale = [
        [0.0, "#1F2937"],
        [0.5, "#FEF3C7"],
        [1.0, "#D1FAE5"],
    ]

    fig_hm = go.Figure(go.Heatmap(
        z=z_vals,
        x=x_labels,
        y=y_labels,
        colorscale=colorscale,
        zmin=0, zmax=2,
        text=[["✓" if v == 2 else ("◐" if v == 1 else "—") for v in row] for row in z_vals],
        texttemplate="%{text}",
        textfont={"size": 14, "color": "white"},
        showscale=False,
        hoverongaps=False,
        hovertemplate="%{y}<br>%{x}<br>%{text}<extra></extra>",
    ))
    fig_hm.update_layout(
        paper_bgcolor="#0B1120",
        plot_bgcolor="#0B1120",
        font=dict(color="#9CA3AF", family="Arial"),
        margin=dict(l=220, r=20, t=30, b=80),
        height=520,
        xaxis=dict(tickfont=dict(size=11, color="#D1D5DB"), side="top"),
        yaxis=dict(tickfont=dict(size=10, color="#9CA3AF"), autorange="reversed"),
    )
    st.plotly_chart(fig_hm, use_container_width=True)

    st.html('<div style="font-size:.7rem;color:#6B7280;margin-top:-10px">✓ = Alignement confirmé (rapports officiels) &nbsp;·&nbsp; ◐ = Alignement indirect / sectoriel &nbsp;·&nbsp; — = Non documenté</div>')

    # Barchart nombre ODD par org
    st.html('<div class="divider"></div>')
    st.html('<span class="section-hdr">ODD COUVERTS PAR ORGANISATION</span>')
    bar_data = {
        "Organisation": [o["nom"] for o in ORGS],
        "Direct": [len(o["odds_direct"]) for o in ORGS],
        "Indirect": [len(o.get("odds_indirect", [])) for o in ORGS],
        "Couleur": [o["couleur"] for o in ORGS],
    }
    fig_bar = go.Figure()
    fig_bar.add_trace(go.Bar(
        name="Direct (confirmé)",
        x=bar_data["Organisation"],
        y=bar_data["Direct"],
        marker_color=[o["couleur"] for o in ORGS],
        text=bar_data["Direct"],
        textposition="inside",
        textfont=dict(color="white", size=13, family="Arial"),
    ))
    fig_bar.add_trace(go.Bar(
        name="Indirect / sectoriel",
        x=bar_data["Organisation"],
        y=bar_data["Indirect"],
        marker_color=["#374151"] * 4,
        text=bar_data["Indirect"],
        textposition="inside",
        textfont=dict(color="#9CA3AF", size=11, family="Arial"),
    ))
    fig_bar.update_layout(
        barmode="stack",
        paper_bgcolor="#0B1120",
        plot_bgcolor="#0B1120",
        font=dict(color="#9CA3AF", family="Arial"),
        legend=dict(
            bgcolor="#111827", bordercolor="#1F2937", borderwidth=1,
            font=dict(color="#D1D5DB", size=11),
        ),
        margin=dict(l=20, r=20, t=10, b=20),
        height=280,
        xaxis=dict(gridcolor="#1F2937", tickfont=dict(color="#D1D5DB")),
        yaxis=dict(gridcolor="#1F2937", tickfont=dict(color="#6B7280"), title="Nombre ODD"),
    )
    st.plotly_chart(fig_bar, use_container_width=True)

# ══ TAB 3 — Comparatif ESG ═══════════════════════════════════════════════════
with tab3:
    st.html('<span class="section-hdr">COMPARATIF ENGAGEMENTS ESG</span>')

    col_a, col_b = st.columns(2)

    # Radar multi-org
    with col_a:
        st.html('<div style="font-size:.75rem;color:#9CA3AF;margin-bottom:6px">PROFIL ESG PAR ORGANISATION</div>')
        cats = ["Environnement", "Social", "Gouvernance"]
        fig_multi = go.Figure()
        for org in ORGS:
            scores = [len(org["esg"]["E"]), len(org["esg"]["S"]), len(org["esg"]["G"])]
            fig_multi.add_trace(go.Scatterpolar(
                r=scores + [scores[0]],
                theta=cats + [cats[0]],
                name=org["court"],
                line_color=org["couleur"],
                fillcolor=_hex_rgba(org["couleur"], 0.13),
                fill="toself",
            ))
        fig_multi.update_layout(
            polar=dict(
                bgcolor="#0B1120",
                radialaxis=dict(visible=True, range=[0, 5], color="#6B7280", gridcolor="#1F2937"),
                angularaxis=dict(color="#9CA3AF", gridcolor="#1F2937"),
            ),
            paper_bgcolor="#111827",
            legend=dict(bgcolor="#111827", font=dict(color="#D1D5DB", size=10)),
            margin=dict(l=40, r=40, t=40, b=40),
            height=320,
            font=dict(family="Arial"),
        )
        st.plotly_chart(fig_multi, use_container_width=True)

    # Actifs sous gestion
    with col_b:
        st.html('<div style="font-size:.75rem;color:#9CA3AF;margin-bottom:6px">ACTIFS GÉRÉS (G$)</div>')
        aum_vals = [7.5, 21.9, 4.9, 0.316]
        aum_labels = ["IQ", "FSTQ", "Desjardins Capital", "DEC"]
        aum_colors = [o["couleur"] for o in ORGS]
        fig_pie = go.Figure(go.Pie(
            labels=aum_labels,
            values=aum_vals,
            marker_colors=aum_colors,
            textinfo="label+percent",
            textfont=dict(size=12, color="white", family="Arial"),
            hole=0.45,
            hovertemplate="%{label}<br>%{value} G$<br>%{percent}<extra></extra>",
        ))
        fig_pie.add_annotation(
            text="~34,6 G$<br><span style='font-size:10px;color:#6B7280'>total</span>",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=13, color="#34D399", family="Arial"),
        )
        fig_pie.update_layout(
            paper_bgcolor="#111827",
            font=dict(family="Arial", color="#9CA3AF"),
            legend=dict(bgcolor="#111827", font=dict(color="#D1D5DB", size=10)),
            margin=dict(l=10, r=10, t=10, b=10),
            height=320,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    # Tableau engagements ESG détaillés
    st.html('<div class="divider"></div>')
    st.html('<span class="section-hdr">ENGAGEMENTS DÉTAILLÉS</span>')

    esg_flat = []
    for org in ORGS:
        for cat, items in org["esg"].items():
            for item in items:
                esg_flat.append({
                    "Organisation": org["court"],
                    "Catégorie": cat,
                    "Engagement": item,
                })
    df_esg = pd.DataFrame(esg_flat)

    cat_filter = st.multiselect(
        "Filtrer par catégorie ESG",
        options=["Environnement", "Social", "Gouvernance"],
        default=["Environnement", "Social", "Gouvernance"],
        key="cat_filter",
    )
    org_filter = st.multiselect(
        "Filtrer par organisation",
        options=[o["court"] for o in ORGS],
        default=[o["court"] for o in ORGS],
        key="org_filter",
    )
    df_show = df_esg[df_esg["Catégorie"].isin(cat_filter) & df_esg["Organisation"].isin(org_filter)]
    st.dataframe(
        df_show.reset_index(drop=True),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Organisation": st.column_config.TextColumn(width=140),
            "Catégorie":    st.column_config.TextColumn(width=120),
            "Engagement":   st.column_config.TextColumn(width=600),
        },
    )

# ══ TAB 4 — Données brutes ════════════════════════════════════════════════════
with tab4:
    st.html('<span class="section-hdr">DONNÉES DE RÉFÉRENCE</span>')

    df_raw = pd.DataFrame([{
        "Acronyme":     o["court"],
        "Organisation": o["nom"],
        "Type":         o["type"],
        "Ville":        o["ville"],
        "Fondation":    o["fondation"],
        "Actif géré":   o["aum"],
        "NEQ":          o["neq"],
        "Site web":     o["web"],
        "ODD directs":  ", ".join(o["odds_direct"]),
        "Nb ODD directs": len(o["odds_direct"]),
    } for o in ORGS])

    st.dataframe(df_raw, use_container_width=True, hide_index=True)

    st.html('<div class="note-box">⚠️ <b>Note sur les NEQ :</b> Les numéros d\'entreprise du Québec (NEQ) de IQ, FSTQ et CRCD sont disponibles sur <b>registreentreprises.gouv.qc.ca</b> (recherche par nom exact). DEC est une agence fédérale non assujettie au registre provincial.</div>')
    st.html('<div style="margin-top:12px;font-size:.68rem;color:#4B5563">Sources : Registre des entreprises du Québec · investquebec.com · fondsftq.com · capitalregional.com · desjardins.com · dec.canada.ca · Rapports ESG/DD 2024-2025</div>')
